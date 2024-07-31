import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from tkcalendar import Calendar
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import os

def browse_excel_file(entry_widget):
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    entry_widget.delete(0, tk.END)
    entry_widget.insert(0, file_path)

def get_selected_dates():
    # start_date = cal_start.get_date()
    # start_date_list = start_date.split('/')
    # select_start_date = start_date_list[1] + "-" + start_date_list[0] + "-" + start_date_list[2]
    # end_date = cal_end.get_date()
    # end_date_list = end_date.split('/')
    # select_end_date = end_date_list[1] + "-" + end_date_list[0] + "-" + end_date_list[2]

    start_date = cal_start.selection_get()
    select_start_date = start_date.strftime("%d-%m-%Y")

    end_date = cal_end.selection_get()
    select_end_date = end_date.strftime("%d-%m-%Y")

    date.config(text="Selected Date from: " + select_start_date + " to " + select_end_date)

def save_result_file():
    result_directory = filedialog.askdirectory()

    if result_directory:
        try:
            # Define the name of the result Excel file
            result_file_name = 'PR status - .xlsx'

            # Specify the full path for the result file, combining the directory and file name
            result_file_path = os.path.join(result_directory, result_file_name)
            
            first_excel_path = entry_first_excel.get()
            second_excel_path = entry_second_excel.get()

            usecol_old = ['Inc Id2','Production/UAT Defect/ Dev','Log Date','Closed date ','Dept','Item (Product / Work Product) ID','Type','Severity','Status','Remark']
            usecol = ['Inc Id2','Production/UAT Defect/ Dev','Actual Working Date','Closed date','Dept','Item (Product / Work Product) ID','Type','Severity','Status','Remark']

            first_excel = pd.read_excel(first_excel_path, usecols=usecol, engine='openpyxl', sheet_name='Defect Log')
            second_excel = pd.read_excel(second_excel_path, usecols=usecol_old, engine='openpyxl', sheet_name='PR Activities')

            start_date = cal_start.get_date()
            end_date = cal_end.get_date()

            first_excel['Actual Working Date'] = pd.to_datetime(first_excel['Actual Working Date'])

            excel1 = second_excel[second_excel['Status'].isin(['open', 'Open'])]
            num_excel1 = len(excel1)
            excel2 = first_excel[(first_excel['Actual Working Date'] >= start_date) & (first_excel['Actual Working Date'] <= end_date) & (first_excel['Type'].isin(['support','Support','PR development' ,'PR Development', 'Data Sharing', 'Defect','defect', 'Service Request','service request']))]
            num_excel2 = len(excel2)
            merged_excel = pd.concat([excel1, excel2], ignore_index=True)
            merged_excel.to_excel(result_file_path, index=False, engine='openpyxl',sheet_name='PR Activities')

            df_new = pd.read_excel(result_file_path, engine='openpyxl', sheet_name='PR Activities')

            filtered_df = df_new[(df_new['Status'].isin(['Closed','closed'])) & (~df_new['Type'].isin(['Defect','defect'])) & (~df_new['Type'].isin(['service request', 'Service Request']))]
            num_excel3 = len(filtered_df)

            filtered_df = df_new[(df_new['Status'].isin(['open', 'Open'])) & (~df_new['Type'].isin(['Defect','defect'])) & (~df_new['Type'].isin(['service request', 'Service Request']))]
            num_excel4 = len(filtered_df)

            filtered_df = df_new[(df_new['Status'].isin(['Closed','closed'])) & (df_new['Type'].isin(['Defect','defect']))]
            Defect_closed = len(filtered_df)

            filtered_df = df_new[(df_new['Status'].isin(['open', 'Open'])) & (df_new['Type'].isin(['Defect','defect']))]
            Defect_open = len(filtered_df)

            filtered_df = df_new[(df_new['Status'].isin(['open', 'Open'])) & (df_new['Type'].isin(['service request', 'Service Request']))]
            Sr_open = len(filtered_df)

            filtered_df = df_new[(df_new['Status'].isin(['Closed','closed'])) & (df_new['Type'].isin(['service request', 'Service Request']))]
            Sr_closed = len(filtered_df)

            new_sheet_name = 'Dashboard'
            file_path = result_file_path

            new_sheet_data = pd.DataFrame({
                ' ': ['Open PR,SR incidents from Last  Week', 'PR Incidents Received(current week)','PR Incidents Closed','PR Incidents Open','Service Request Closed','Service Request Open','Defect Closed','Defect Open'],
                'S1': ['','', '','','','','',''],
                'S2': ['','', '','','','','',''],
                'S3': ['','', '','','','','',''],
                'S4': ['','', '','','','','',''],
                'Total': [num_excel1,num_excel2, num_excel3,num_excel4,Sr_closed,Sr_open,Defect_closed,Defect_open],
            })

            try:
                workbook = load_workbook(file_path)
                sheet1 = workbook['PR Activities']
                if new_sheet_name not in workbook.sheetnames:
                    new_sheet = workbook.create_sheet(title=new_sheet_name)
                else:
                    new_sheet = workbook[new_sheet_name]
                start_row_insert = 9
                start_col_insert = 12

                for row_index, row in enumerate(dataframe_to_rows(new_sheet_data, index=False, header=True),
                                                start=start_row_insert):
                    for col_index, value in enumerate(row, start=start_col_insert):
                        new_sheet.cell(row=row_index, column=col_index, value=value)

                all_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                    bottom=Side(style='thin'))

                for row in new_sheet.iter_rows(min_row=start_row_insert, max_row=new_sheet.max_row,
                                                min_col=start_col_insert, max_col=new_sheet.max_column):
                    for cell in row:
                        cell.border = all_border

                for row in sheet1.iter_rows(min_row=1, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
                    for cell in row:
                        cell.border = all_border

                workbook.save(file_path)

                date.config(text="File has been saved at your selected location : " + file_path)

            except FileNotFoundError:
                print(f"File '{file_path}' not found.")
            except Exception as e:
                print(f"An error occurred: {str(e)}")
        except Exception as e:
            print(f"An error occurred while saving the result file: {str(e)}")




# Create the main application window
root = tk.Tk()
root.title("Weekly Status Report")

file_frame = tk.LabelFrame(root, text="Excel Files")
file_frame.pack(pady=10, padx=10, fill="both", expand="yes")

label_first_excel = tk.Label(file_frame, text="Select First Excel File:")
label_first_excel.grid(row=0, column=0, padx=5, pady=5)
entry_first_excel = tk.Entry(file_frame)
entry_first_excel.grid(row=0, column=1, padx=5, pady=5)
browse_button_first_excel = tk.Button(file_frame, text="Browse", command=lambda: browse_excel_file(entry_first_excel))
browse_button_first_excel.grid(row=0, column=2, padx=5, pady=5)

label_second_excel = tk.Label(file_frame, text="Select Second Excel File:")
label_second_excel.grid(row=0, column=5, padx=5, pady=5)
entry_second_excel = tk.Entry(file_frame)
entry_second_excel.grid(row=0, column=6, padx=5, pady=5)
browse_button_second_excel = tk.Button(file_frame, text="Browse", command=lambda: browse_excel_file(entry_second_excel))
browse_button_second_excel.grid(row=0, column=7, padx=5, pady=5)

date_frame = tk.LabelFrame(root, text="Date Selection")
date_frame.pack(pady=10, padx=10, fill="both", expand="yes")

label_start_date = tk.Label(date_frame, text="Select Start Date:")
label_start_date.grid(row=0, column=0, padx=5, pady=5)
cal_start = Calendar(date_frame)
cal_start.grid(row=0, column=1, padx=5, pady=5)

label_end_date = tk.Label(date_frame, text="Select End Date:")
label_end_date.grid(row=0, column=2, padx=5, pady=5)
cal_end = Calendar(date_frame)
cal_end.grid(row=0, column=3, padx=5, pady=5)

button_frame = tk.Frame(root)
button_frame.pack(pady=10, padx=10, fill="both", expand="yes")

get_dates_button = tk.Button(button_frame, text="Get Selected Dates", command=get_selected_dates)
get_dates_button.grid(row=0, column=0, padx=5, pady=5)

save_button = tk.Button(button_frame, text="Save Result", command=save_result_file)
save_button.grid(row=0, column=2, padx=5, pady=5)

exit_button = tk.Button(button_frame, text="Exit", command=root.destroy)
exit_button.grid(row=0, column=3, padx=5, pady=5)

date = tk.Label(date_frame, text="")
date.grid(row=5, column=1, padx=5, pady=5)

root.mainloop()